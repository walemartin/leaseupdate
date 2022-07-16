#!/usr/bin/env python
#import pandas as pd
import openpyxl
#from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string
from openpyxl.utils import get_column_letter
import smtplib
import ssl
#from email.message import EmailMessage
from openpyxl import Workbook, load_workbook
#from openpyxl.utils import get_column_letter
#from openpyxl.styles import Font


# Give the location of the file
path = "leasebook.xlsx"
 
# workbook object is created
wb_obj = openpyxl.load_workbook(path)


sheet_obj = wb_obj.active
# cell references (original spreadsheet) 
min_column = wb_obj.active.min_column
max_column = wb_obj.active.max_column
min_row = wb_obj.active.min_row
max_row = wb_obj.active.max_row

print(min_column,max_column,min_row,max_row)
sheet_obj.auto_filter.ref="A1:H1"

for rowNum in range(2,sheet_obj.max_row+1):
	n='=DATEDIF(D{},E{},"D")'.format(rowNum,rowNum)
	m ='=IF(TODAY()>E{},"Rent Expired","Active")'.format(rowNum,rowNum)
	sheet_obj.cell(row = rowNum, column = 6).value = n
	sheet_obj.cell(row=rowNum,column=7).value='=E{}-TODAY()'.format(rowNum,rowNum)
	sheet_obj.cell(row=rowNum,column=8).value=m
	#sheet_obj.cell(row = rowNum, column = 6).value ='=DATEDIF(D2,E2,"D")'


wb_obj.save(path)
wb_obj.close()
print("Excel sheet data updated successfully")



# Give the location of the file
path2 = "NewGrades.xlsx"
 
# workbook object is created
wb = openpyxl.load_workbook(path2)

data = {
	"Joe": {
		"math": 65,
		"science": 78,
		"english": 98,
		"gym": 89
	},
	"Bill": {
		"math": 55,
		"science": 72,
		"english": 87,
		"gym": 95
	},
	"Tim": {
		"math": 100,
		"science": 45,
		"english": 75,
		"gym": 92
	},
	"Sally": {
		"math": 30,
		"science": 25,
		"english": 45,
		"gym": 100
	},
	"Jane": {
		"math": 100,
		"science": 100,
		"english": 100,
		"gym": 60
	}
}

#wb = Workbook()
ws = wb.active
ws.title = "Grades"

headings = ['Name'] + list(data['Joe'].keys())
ws.append(headings)

for person in data:
	grades = list(data[person].values())
	ws.append([person] + grades)

for col in range(2, len(data['Joe']) + 2):
	char = get_column_letter(col)
	ws[char + "7"] = f"=SUM({char + '2'}:{char + '6'})/{len(data)}"

for col in range(1, 6):
	ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save(path2)


#df = pd.read_excel('C:\\Users\\OLAWALE MARTINS\\Documents\\leasebook.xlsx')
#print(df.head())

wb.close()
print("Another excel file data updated successfully")





